"""
core/export_service.py
=====================
Data export service supporting CSV, Excel, and PDF formats.

Provides utilities for exporting audit logs, compliance reports, and other data.
"""

import csv
import io
import logging
from datetime import datetime
from django.http import HttpResponse

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

logger = logging.getLogger(__name__)


class ExportService:
    """Service for exporting data to various formats."""

    @staticmethod
    def export_to_csv(data, filename="export.csv", headers=None):
        """
        Export data to CSV format.

        Args:
            data: List of dictionaries or QuerySet
            filename: Output filename
            headers: List of header names (if None, uses dict keys)

        Returns:
            HttpResponse: CSV file response
        """
        # Convert QuerySet to list of dicts if needed
        if hasattr(data, 'values'):
            data = list(data.values())

        if not data:
            data = [{}]

        # Get headers
        if headers is None and data:
            headers = list(data[0].keys())

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        writer = csv.DictWriter(response, fieldnames=headers)
        writer.writeheader()

        for row in data:
            # Convert datetime objects to strings
            clean_row = {}
            for key, value in row.items():
                if isinstance(value, datetime):
                    clean_row[key] = value.isoformat()
                elif isinstance(value, (list, dict)):
                    clean_row[key] = str(value)
                else:
                    clean_row[key] = value or ''

            writer.writerow(clean_row)

        return response

    @staticmethod
    def export_to_excel(data, filename="export.xlsx", headers=None, title=None, include_timestamp=True):
        """
        Export data to Excel format with formatting.

        Args:
            data: List of dictionaries or QuerySet
            filename: Output filename
            headers: List of header names
            title: Worksheet title
            include_timestamp: Add timestamp to workbook

        Returns:
            HttpResponse: Excel file response
        """
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl not installed")

        # Convert QuerySet to list
        if hasattr(data, 'values'):
            data = list(data.values())

        if not data:
            data = [{}]

        # Get headers
        if headers is None and data:
            headers = list(data[0].keys())

        # Create workbook
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = title or "Data"

        # Add title if provided
        if title:
            worksheet['A1'] = title
            worksheet['A1'].font = Font(size=14, bold=True)
            worksheet.merge_cells('A1:Z1')
            start_row = 3
        else:
            start_row = 1

        # Add timestamp
        if include_timestamp:
            timestamp_row = start_row - 1
            worksheet[f'A{timestamp_row}'] = f"Generated: {datetime.now().isoformat()}"
            worksheet[f'A{timestamp_row}'].font = Font(size=10, italic=True)

        # Write headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col_idx, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=start_row, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Write data
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row_idx, row_data in enumerate(data, start=start_row + 1):
            for col_idx, header in enumerate(headers, start=1):
                value = row_data.get(header, '')

                # Format datetime
                if isinstance(value, datetime):
                    value = value.isoformat()
                elif isinstance(value, (list, dict)):
                    value = str(value)

                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        # Adjust column widths
        for col_idx, header in enumerate(headers, start=1):
            width = min(max(len(str(header)), 10), 50)
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

        # Freeze header row
        worksheet.freeze_panes = f'A{start_row + 1}'

        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        workbook.save(response)
        return response

    @staticmethod
    def export_to_pdf(data, filename="export.pdf", title="Export Report", headers=None):
        """
        Export data to PDF format with table.

        Args:
            data: List of dictionaries or QuerySet
            filename: Output filename
            title: Report title
            headers: Column headers

        Returns:
            HttpResponse: PDF file response
        """
        if not PDF_AVAILABLE:
            raise ImportError("reportlab not installed")

        # Convert QuerySet to list
        if hasattr(data, 'values'):
            data = list(data.values())

        if not data:
            return HttpResponse("No data to export", content_type='text/plain')

        # Get headers
        if headers is None and data:
            headers = list(data[0].keys())

        # Create PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)

        # Build content
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#1f4788'),
            spaceAfter=30,
            alignment=1,  # center
        )

        story = []

        # Add title
        story.append(Paragraph(title, title_style))
        story.append(Spacer(1, 0.2 * inch))

        # Add timestamp
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        story.append(Paragraph(f"<i>Generated: {timestamp}</i>", styles['Normal']))
        story.append(Spacer(1, 0.2 * inch))

        # Prepare table data
        table_data = [headers]

        for row in data:
            row_data = []
            for header in headers:
                value = row.get(header, '')

                # Format datetime
                if isinstance(value, datetime):
                    value = value.strftime("%Y-%m-%d %H:%M:%S")
                elif isinstance(value, (list, dict)):
                    value = str(value)[:100]  # Truncate long values

                row_data.append(str(value or ''))

            table_data.append(row_data)

        # Create table
        if len(table_data) > 1:
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0F0F0')]),
            ]))

            story.append(table)

        # Build PDF
        doc.build(story)

        # Create response
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response.write(buffer.getvalue())

        return response

    @staticmethod
    def get_export_formats():
        """Get list of available export formats."""
        formats = ['csv']

        if EXCEL_AVAILABLE:
            formats.append('excel')

        if PDF_AVAILABLE:
            formats.append('pdf')

        return formats


def export_queryset(queryset, format='csv', filename=None, headers=None, **kwargs):
    """
    Convenience function to export a QuerySet.

    Args:
        queryset: Django QuerySet
        format: 'csv', 'excel', or 'pdf'
        filename: Output filename
        headers: Column headers
        **kwargs: Additional arguments

    Returns:
        HttpResponse: File response
    """
    if format == 'csv':
        return ExportService.export_to_csv(queryset, filename or 'export.csv', headers)
    elif format == 'excel':
        return ExportService.export_to_excel(queryset, filename or 'export.xlsx', headers, **kwargs)
    elif format == 'pdf':
        return ExportService.export_to_pdf(queryset, filename or 'export.pdf', headers=headers, **kwargs)
    else:
        raise ValueError(f"Unsupported format: {format}")
